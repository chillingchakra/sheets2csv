#!/usr/bin/env python

# export data sheets from xlsx to csv

from openpyxl import load_workbook
import csv
from os import sys
import os

def get_all_sheets(excel_file):
    sheets = []
    workbook = load_workbook(excel_file) #,use_iterators=True,data_only=True)
    all_worksheets = workbook.get_sheet_names()
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    return sheets

def csv_from_excel(excel_file, sheets):
    workbook = load_workbook(excel_file,data_only=True)
    for worksheet_name in sheets:
        print("Export " + worksheet_name + " ...")

        try:
            worksheet = workbook.get_sheet_by_name(worksheet_name)
        except KeyError:
            print("Could not find " + worksheet_name)
        
            sys.exit(1)
        try:    
                wname = excel_file.split('/')[1].split('.')[0] +'-'+ worksheet_name.split('size')[1].replace('#','').strip()
        except:
                   print 'error', excel_file,worksheet_name


        your_csv_file = open(''.join(['csv-expo','/',wname,'.csv']), 'wb')
        wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)
        for row in worksheet.iter_rows():
            lrow = []
            for cell in row:
                lrow.append(cell.value)
            wr.writerow(lrow)
        print(" ... done")
	your_csv_file.close()
path='xlsx-in'

allfiles = os.listdir(path)


#Attempting Deletion of Odd Windows Files that Are
#Generated which also have xlsx extension"

#try:
#    allfiles = allfiles.find("size") != -1
#except KeyError:
#    print("Did not match our criterie")

for onefile in allfiles:
    if onefile.startswith("~$"):
        os.remove(os.path.join(onefile))
    if onefile.endswith(".xlsx"):    
        sheets = get_all_sheets(path+'/'+onefile)    
        csv_from_excel(path+'/'+onefile, sheets)
#if not 2 <= len(sys.argv) <= 3:
#	print("Call with " + sys.argv[0] + " <xlxs file> [comma separated list of sheets to export]")
#	sys.exit(1)
#lse:
#    sheets = []
#    if len(sys.argv) == 3:
#3        sheets = list(sys.argv[2].split(','))
#    else:
#        sheets = get_all_sheets(sys.argv[1])
#    assert(sheets != None and len(sheets) > 0)
#    csv_from_excel(sys.argv[1], sheets)


